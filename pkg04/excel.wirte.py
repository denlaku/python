# coding=UTF-8
# pylint: disable=invalid-name
'''
read excel
'''
from openpyxl import Workbook
from openpyxl.compat import range as orange
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'pkg03/empty_book.xlsx'
ws1 = wb.active  # 第一个表
ws1.title = "range names"  # 第一个表命名
# 遍历第一个表的1到40行，赋值一个600内的随机数
for row in orange(1, 3):
    ws1.append(orange(5))
ws2 = wb.create_sheet(title="Pi")
ws2['D5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in orange(2, 4):
    for col in orange(3, 6):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))

wb.save(filename=dest_filename)

print "-------------------------"
